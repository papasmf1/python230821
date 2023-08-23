from openpyxl import Workbook
from random import randint
from datetime import datetime, timedelta

# Create a new Excel workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active

# Write headers
headers = ["Product Name", "Price", "Quantity", "Sale Date"]
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Generate and write sales data
product_names = ["Laptop", "Smartphone", "Tablet", "Headphones", "Smartwatch"]
start_date = datetime(2023, 1, 1)
for row_num in range(2, 102):
    product_name = product_names[randint(0, len(product_names) - 1)]
    price = round(randint(300, 1500) + randint(1, 99) / 100, 2)
    quantity = randint(1, 10)
    sale_date = start_date + timedelta(days=randint(0, 240))
    
    data = [product_name, price, quantity, sale_date.strftime('%Y-%m-%d')]
    for col_num, value in enumerate(data, 1):
        sheet.cell(row=row_num, column=col_num, value=value)

# Save the workbook to the specified file path
file_path = "c:\\work\\sales.xlsx"
workbook.save(file_path)

print(f"Sales data has been saved to {file_path}")
